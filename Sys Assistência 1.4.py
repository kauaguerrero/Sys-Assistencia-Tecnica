from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import json
from datetime import datetime

clientes = []
vendas = []
NOME_ARQUIVO_EXCEL = "vendas.xlsx"
ARQUIVO_CLIENTES = "clientes.json"

def carregar_planilha():
    if os.path.exists(NOME_ARQUIVO_EXCEL):
        wb = load_workbook(NOME_ARQUIVO_EXCEL)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendas"
        ws.append(["Produto", "Valor (R$)", "Cliente", "CPF", "Data e Hora"])
    return wb, ws

def ajustar_largura_colunas(ws):
    for coluna in ws.columns:
        max_length = 0
        coluna_letra = get_column_letter(coluna[0].column)
        for cell in coluna:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[coluna_letra].width = max_length + 2

def carregar_clientes():
    if os.path.exists(ARQUIVO_CLIENTES):
        with open(ARQUIVO_CLIENTES, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def salvar_clientes():
    with open(ARQUIVO_CLIENTES, "w", encoding="utf-8") as f:
        json.dump(clientes, f, ensure_ascii=False, indent=4)

clientes = carregar_clientes()

def validar_cpf(cpf):
    return cpf.isdigit() and len(cpf) == 11

def formatar_moeda(valor):
    return f"R${float(valor):.2f}"

def criar_cliente():
    nome = input('Digite o nome do cliente: ').strip()
    cpf = input('Digite o CPF do cliente (apenas números): ').strip()
    telefone = input('Digite o telefone do cliente: ').strip()

    if not validar_cpf(cpf):
        print("❌ CPF inválido. Deve conter 11 dígitos numéricos.")
        return

    cliente = {
        'Nome': nome,
        'CPF': cpf,
        'Telefone': telefone
    }
    clientes.append(cliente)
    salvar_clientes()
    print('✅ Cliente cadastrado com sucesso!')

def exibir_clientes():
    if not clientes:
        print('⚠️ Nenhum cliente cadastrado.')
        return

    print('\n📋 Lista de Clientes:')
    for i, cliente in enumerate(clientes, 1):
        print(f"{i}. Nome: {cliente['Nome']} | CPF: {cliente['CPF']} | Telefone: {cliente['Telefone']}")

def remover_cliente():
    if not clientes:
        print('⚠️ Nenhum cliente para remover.')
        return

    exibir_clientes()
    try:
        indice = int(input('\nDigite o número do cliente a remover: ')) - 1
        if 0 <= indice < len(clientes):
            removido = clientes.pop(indice)
            salvar_clientes()
            print(f"✅ Cliente '{removido['Nome']}' removido com sucesso.")
        else:
            print('❌ Índice inválido.')
    except ValueError:
        print('❌ Entrada inválida. Digite um número.')

def registrar_venda():
    if not clientes:
        print('⚠️ Nenhum cliente cadastrado para venda.')
        return

    exibir_clientes()
    try:
        indice = int(input('\nDigite o número do cliente: ')) - 1
        if not (0 <= indice < len(clientes)):
            print('❌ Cliente inválido.')
            return

        cliente = clientes[indice]
        produto = input('Digite o nome do produto/serviço: ').strip().upper()
        valor = float(input('Digite o valor (R$): '))

        venda = {
            'Produto': produto,
            'Valor': valor,
            'Cliente': cliente['Nome'],
            'CPF': cliente['CPF'],
            'DataHora': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        }
        vendas.append(venda)
        salvar_venda(venda)
        print('✅ Venda registrada com sucesso!')

    except ValueError:
        print('❌ Valor inválido. Use números para valor e índice.')

def salvar_venda(venda):
    try:
        wb, ws = carregar_planilha()
        ws.append([
            venda['Produto'],
            venda['Valor'],
            venda['Cliente'],
            venda['CPF'],
            venda['DataHora']
        ])
        ajustar_largura_colunas(ws)
        wb.save(NOME_ARQUIVO_EXCEL)
    except PermissionError:
        print("❌ Erro: O arquivo Excel está aberto. Feche-o para salvar.")
    except Exception as e:
        print(f"❌ Erro ao salvar: {str(e)}")

def listar_vendas():
    if not vendas:
        print('⚠️ Nenhuma venda pendente.')
        return

    print('\n🛒 Vendas Registradas:')
    for i, venda in enumerate(vendas, 1):
        print(
            f"{i}. {venda['Produto']} | "
            f"{formatar_moeda(venda['Valor'])} | "
            f"Cliente: {venda['Cliente']} | "
            f"CPF: {venda['CPF']}"
        )


def limpar_planilha():
    try:
        wb, ws = carregar_planilha()

        ws.delete_rows(2, ws.max_row)

        wb.save(NOME_ARQUIVO_EXCEL)
        print("✅ Planilha limpa com sucesso! Apenas os cabeçalhos foram mantidos.")

    except PermissionError:
        print("❌ Erro: O arquivo Excel está aberto. Feche-o para limpar.")
    except Exception as e:
        print(f"❌ Erro ao limpar planilha: {str(e)}")

def menu():
    while True:
        print("\n" + "=" * 30)
        print("MENU PRINCIPAL".center(30))
        print("=" * 30)
        print("1. Cadastrar Cliente")
        print("2. Listar Clientes")
        print("3. Remover Cliente")
        print("4. Registrar Venda")
        print("5. Listar Vendas")
        print("6. Limpar planilha de vendas")
        print("0. Sair")
        print("=" * 30)

        escolha = input("Digite sua opção: ").strip()

        match escolha:
            case '1': criar_cliente()
            case '2': exibir_clientes()
            case '3': remover_cliente()
            case '4': registrar_venda()
            case '5': listar_vendas()
            case '6': limpar_planilha()
            case '0':
                print("Saindo do sistema...")
                break
            case _: print("❌ Opção inválida!")

if __name__ == "__main__":
    menu()