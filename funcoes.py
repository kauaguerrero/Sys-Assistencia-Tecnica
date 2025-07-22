import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText
from PIL import Image, ImageTk
import os
import json
from datetime import datetime
import uuid
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Funções auxiliares
def formatar_moeda(valor):
    return f"R${float(valor):.2f}"

def carregar_planilha_vendas():
    if os.path.exists(NOME_ARQUIVO_VENDAS):
        wb = load_workbook(NOME_ARQUIVO_VENDAS)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendas"
        ws.append(["Tipo", "Descrição", "Valor (R$)", "Cliente", "CPF", "Data e Hora", "Código Produto", "Quantidade", "ID Venda"])
    return wb, ws

def ajustar_largura_colunas(ws):
    for coluna in ws.columns:
        max_length = 0
        coluna_letra = get_column_letter(coluna[0].column)
        for cell in coluna:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[coluna_letra].width = max_length + 2

def carregar_dados():
    global clientes, estoque

    # Carregar clientes
    if os.path.exists(ARQUIVO_CLIENTES):
        with open(ARQUIVO_CLIENTES, "r", encoding="utf-8") as f:
            clientes = json.load(f)

    # Carregar estoque
    if os.path.exists(ARQUIVO_ESTOQUE):
        with open(ARQUIVO_ESTOQUE, "r", encoding="utf-8") as f:
            estoque = json.load(f)

def salvar_clientes():
    with open(ARQUIVO_CLIENTES, "w", encoding="utf-8") as f:
        json.dump(clientes, f, ensure_ascii=False, indent=4)

def salvar_estoque():
    with open(ARQUIVO_ESTOQUE, "w", encoding="utf-8") as f:
        json.dump(estoque, f, ensure_ascii=False, indent=4)

def validar_cpf(cpf):
    return len(cpf) == 11 and cpf.isdigit()

def gerar_codigo_produto():
    return str(uuid.uuid4())[:4].upper()

def fechamento_caixa_diario():
    try:
        wb, ws = carregar_planilha_vendas()

        if ws.max_row == 1:
            messagebox.showinfo("Informação", "Nenhuma venda registrada hoje.")
            return

        data_hoje = datetime.now().strftime('%d/%m/%Y')
        vendas_hoje = []
        total_produtos = 0.0
        total_servicos = 0.0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[5].startswith(data_hoje):
                vendas_hoje.append(row)
                if row[0] == "PRODUTO":
                    total_produtos += row[2]
                elif row[0] == "SERVIÇO":
                    total_servicos += row[2]

        if not vendas_hoje:
            messagebox.showinfo("Informação", f"Nenhuma venda registrada hoje ({data_hoje}).")
            return

        total_geral = total_produtos + total_servicos
        quantidade_vendas = len(vendas_hoje)

        data_arquivo = datetime.now().strftime('%Y%m%d')
        nome_arquivo = f"fechamento_{data_arquivo}.xlsx"
        wb_fechamento = Workbook()
        ws_fechamento = wb_fechamento.active
        ws_fechamento.title = "Fechamento Diário"

        ws_fechamento.append(["FECHAMENTO DIÁRIO"])
        ws_fechamento.append(["Data", data_hoje])
        ws_fechamento.append([])
        ws_fechamento.append(["Resumo Financeiro"])
        ws_fechamento.append(["Tipo", "Total (R$)"])
        ws_fechamento.append(["Produtos", total_produtos])
        ws_fechamento.append(["Serviços", total_servicos])
        ws_fechamento.append(["TOTAL GERAL", total_geral])
        ws_fechamento.append([])
        ws_fechamento.append(["Detalhamento de Vendas"])
        ws_fechamento.append(["Tipo", "Descrição", "Valor (R$)", "Cliente", "CPF", "Hora", "Código", "Quantidade"])

        for venda in vendas_hoje:
            ws_fechamento.append(venda)

        ajustar_largura_colunas(ws_fechamento)
        wb_fechamento.save(nome_arquivo)

        messagebox.showinfo("Sucesso", f"Fechamento diário gerado em {nome_arquivo}")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao gerar fechamento: {str(e)}")

def limpar_historico_vendas():
    try:
        wb, ws = carregar_planilha_vendas()
        ws.delete_rows(2, ws.max_row)
        wb.save(NOME_ARQUIVO_VENDAS)
        messagebox.showinfo("Sucesso", "Histórico de vendas limpo com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao limpar histórico: {str(e)}")

# Variáveis globais
clientes = []
vendas = []
estoque = []
NOME_ARQUIVO_VENDAS = "vendas.xlsx"
ARQUIVO_CLIENTES = "clientes.json"
ARQUIVO_ESTOQUE = "estoque.json"

# Configuração da interface principal
class SistemaVendasApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Vendas")
        self.root.geometry("1200x700")
        self.root.minsize(1000, 600)

        def carregar_dados():
            global clientes, estoque

            # Carregar clientes
            if os.path.exists(ARQUIVO_CLIENTES):
                with open(ARQUIVO_CLIENTES, "r", encoding="utf-8") as f:
                    clientes = json.load(f)

            # Carregar estoque
            if os.path.exists(ARQUIVO_ESTOQUE):
                with open(ARQUIVO_ESTOQUE, "r", encoding="utf-8") as f:
                    estoque = json.load(f)

        def salvar_estoque():
            with open(ARQUIVO_ESTOQUE, "w", encoding="utf-8") as f:
                json.dump(estoque, f, ensure_ascii=False, indent=4)

        def validar_cpf(cpf):
            return len(cpf) == 11 and cpf.isdigit()



        def gerar_codigo_produto():
            return str(uuid.uuid4())[:4].upper()

        # Carregar dados
        carregar_dados()

        # Configurar estilo
        self.setup_style()

        # Criar layout principal
        self.setup_ui()

    def setup_style(self):
        self.style = ttk.Style()
        self.style.theme_use('clam')

        # Cores
        self.bg_color = "#f0f0f0"
        self.primary_color = "#4a6fa5"
        self.secondary_color = "#166088"
        self.accent_color = "#4fc3f7"
        self.danger_color = "#e53935"

        # Configurar estilos
        self.style.configure('TFrame', background=self.bg_color)
        self.style.configure('TLabel', background=self.bg_color, font=('Helvetica', 10))
        self.style.configure('TButton', font=('Helvetica', 10), padding=6)
        self.style.configure('Header.TLabel', font=('Helvetica', 14, 'bold'), foreground=self.primary_color)
        self.style.configure('Primary.TButton', foreground='white', background=self.primary_color)
        self.style.configure('Secondary.TButton', foreground='white', background=self.secondary_color)
        self.style.configure('Accent.TButton', foreground='white', background=self.accent_color)
        self.style.configure('Danger.TButton', foreground='white', background=self.danger_color)
        self.style.map('Primary.TButton', background=[('active', self.secondary_color)])
        self.style.map('Secondary.TButton', background=[('active', self.primary_color)])
        self.style.map('Danger.TButton', background=[('active', '#c62828')])

    def setup_ui(self):
        # Frame principal
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Barra lateral
        self.sidebar = ttk.Frame(self.main_frame, width=200, style='TFrame')
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)

        # Logo
        self.logo_label = ttk.Label(self.sidebar, text="Sistema de Vendas", style='Header.TLabel')
        self.logo_label.pack(pady=20)

        # Botões do menu
        self.menu_buttons = []

        btn_clientes = ttk.Button(self.sidebar, text="Clientes", command=self.show_clientes, style='Primary.TButton')
        btn_clientes.pack(fill=tk.X, pady=5)
        self.menu_buttons.append(btn_clientes)

        btn_estoque = ttk.Button(self.sidebar, text="Estoque", command=self.show_estoque, style='Primary.TButton')
        btn_estoque.pack(fill=tk.X, pady=5)
        self.menu_buttons.append(btn_estoque)

        btn_vendas = ttk.Button(self.sidebar, text="Vendas", command=self.show_vendas, style='Primary.TButton')
        btn_vendas.pack(fill=tk.X, pady=5)
        self.menu_buttons.append(btn_vendas)

        btn_sair = ttk.Button(self.sidebar, text="Sair", command=self.root.quit, style='Danger.TButton')
        btn_sair.pack(fill=tk.X, pady=5, side=tk.BOTTOM)
        self.menu_buttons.append(btn_sair)

        # Área de conteúdo
        self.content_frame = ttk.Frame(self.main_frame, style='TFrame')
        self.content_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Inicializar frames de conteúdo
        self.frames = {}
        for F in (ClientesFrame, EstoqueFrame, VendasFrame):
            frame = F(self.content_frame, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_clientes()

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()

    def show_clientes(self):
        self.show_frame(ClientesFrame)
        self.update_menu_buttons(0)

    def show_estoque(self):
        self.show_frame(EstoqueFrame)
        self.update_menu_buttons(1)

    def show_vendas(self):
        self.show_frame(VendasFrame)
        self.update_menu_buttons(2)

    def update_menu_buttons(self, active_index):
        for i, btn in enumerate(self.menu_buttons[:3]):  # Ignorar o botão Sair
            if i == active_index:
                btn.configure(style='Secondary.TButton')
            else:
                btn.configure(style='Primary.TButton')


# Frame de Clientes
class ClientesFrame(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.setup_ui()

    def setup_ui(self):
        # Frame de cabeçalho
        header_frame = ttk.Frame(self)
        header_frame.pack(fill=tk.X, pady=10)

        ttk.Label(header_frame, text="Gerenciamento de Clientes", style='Header.TLabel').pack(side=tk.LEFT)

        # Frame de botões
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(btn_frame, text="Adicionar Cliente", command=self.adicionar_cliente,
                   style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Remover Cliente", command=self.remover_cliente,
                   style='Danger.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Atualizar Lista", command=self.atualizar_lista,
                   style='Primary.TButton').pack(side=tk.LEFT, padx=5)

        # Lista de clientes
        self.tree_frame = ttk.Frame(self)
        self.tree_frame.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(self.tree_frame, columns=('Nome', 'CPF', 'Telefone'), show='headings')
        self.tree.heading('Nome', text='Nome')
        self.tree.heading('CPF', text='CPF')
        self.tree.heading('Telefone', text='Telefone')

        vsb = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)

        self.atualizar_lista()

    def atualizar_lista(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for cliente in clientes:
            self.tree.insert('', 'end', values=(cliente['Nome'], cliente['CPF'], cliente['Telefone']))

    def adicionar_cliente(self):
        dialog = AdicionarClienteDialog(self)
        self.wait_window(dialog.top)
        self.atualizar_lista()

    def remover_cliente(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Aviso", "Selecione um cliente para remover")
            return

        item = self.tree.item(selected[0])
        nome = item['values'][0]

        if messagebox.askyesno("Confirmar", f"Remover cliente {nome}?"):
            cpf = item['values'][1]
            for i, cliente in enumerate(clientes):
                if cliente['CPF'] == cpf:
                    clientes.pop(i)
                    salvar_clientes()
                    break
            self.atualizar_lista()


# Dialogo para adicionar cliente
class AdicionarClienteDialog:
    def __init__(self, parent):
        self.top = tk.Toplevel(parent)
        self.top.title("Adicionar Cliente")
        self.top.geometry("400x300")

        ttk.Label(self.top, text="Nome:").pack(pady=(10, 0))
        self.nome_entry = ttk.Entry(self.top)
        self.nome_entry.pack(fill=tk.X, padx=20, pady=5)

        ttk.Label(self.top, text="CPF (apenas números):").pack(pady=(10, 0))
        self.cpf_entry = ttk.Entry(self.top)
        self.cpf_entry.pack(fill=tk.X, padx=20, pady=5)

        ttk.Label(self.top, text="Telefone:").pack(pady=(10, 0))
        self.telefone_entry = ttk.Entry(self.top)
        self.telefone_entry.pack(fill=tk.X, padx=20, pady=5)

        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=tk.X, pady=20)

        ttk.Button(btn_frame, text="Cancelar", command=self.top.destroy,
                   style='Danger.TButton').pack(side=tk.LEFT, padx=10, expand=True)
        ttk.Button(btn_frame, text="Salvar", command=self.salvar_cliente,
                   style='Accent.TButton').pack(side=tk.RIGHT, padx=10, expand=True)

    def salvar_cliente(self):
        nome = self.nome_entry.get().strip()
        cpf = self.cpf_entry.get().strip()
        telefone = self.telefone_entry.get().strip()

        if not nome or not cpf or not telefone:
            messagebox.showerror("Erro", "Todos os campos são obrigatórios")
            return

        if not validar_cpf(cpf):
            messagebox.showerror("Erro", "CPF inválido. Deve conter 11 dígitos numéricos.")
            return

        cliente = {
            'Nome': nome,
            'CPF': cpf,
            'Telefone': telefone
        }
        clientes.append(cliente)
        salvar_clientes()
        messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")
        self.top.destroy()


# Frame de Estoque
class EstoqueFrame(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.setup_ui()

    def setup_ui(self):
        # Frame de cabeçalho
        header_frame = ttk.Frame(self)
        header_frame.pack(fill=tk.X, pady=10)

        ttk.Label(header_frame, text="Gerenciamento de Estoque", style='Header.TLabel').pack(side=tk.LEFT)

        # Frame de botões
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(btn_frame, text="Adicionar Produto", command=self.adicionar_produto,
                   style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Atualizar Estoque", command=self.atualizar_estoque,
                   style='Primary.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Remover Produto", command=self.remover_produto,
                   style='Danger.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Atualizar Lista", command=self.atualizar_lista,
                   style='Primary.TButton').pack(side=tk.LEFT, padx=5)

        # Lista de estoque
        self.tree_frame = ttk.Frame(self)
        self.tree_frame.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(self.tree_frame, columns=('Código', 'Nome', 'Preço', 'Quantidade'), show='headings')
        self.tree.heading('Código', text='Código')
        self.tree.heading('Nome', text='Nome')
        self.tree.heading('Preço', text='Preço (R$)')
        self.tree.heading('Quantidade', text='Quantidade')

        vsb = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)

        self.atualizar_lista()

    def atualizar_lista(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for produto in estoque:
            self.tree.insert('', 'end', values=(
                produto['Codigo'],
                produto['Nome'],
                formatar_moeda(produto['Preco']),
                produto['Quantidade']
            ))

    def adicionar_produto(self):
        dialog = AdicionarProdutoDialog(self)
        self.wait_window(dialog.top)
        self.atualizar_lista()

    def atualizar_estoque(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Aviso", "Selecione um produto para atualizar")
            return

        item = self.tree.item(selected[0])
        codigo = item['values'][0]

        dialog = AtualizarEstoqueDialog(self, codigo)
        self.wait_window(dialog.top)
        self.atualizar_lista()

    def remover_produto(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Aviso", "Selecione um produto para remover")
            return

        item = self.tree.item(selected[0])
        nome = item['values'][1]
        codigo = item['values'][0]

        if messagebox.askyesno("Confirmar", f"Remover produto {nome}?"):
            for i, produto in enumerate(estoque):
                if produto['Codigo'] == codigo:
                    estoque.pop(i)
                    salvar_estoque()
                    break
            self.atualizar_lista()


# Diálogos para Estoque
class AdicionarProdutoDialog:
    def __init__(self, parent):
        self.top = tk.Toplevel(parent)
        self.top.title("Adicionar Produto")
        self.top.geometry("400x350")

        ttk.Label(self.top, text="Nome do Produto:").pack(pady=(10, 0))
        self.nome_entry = ttk.Entry(self.top)
        self.nome_entry.pack(fill=tk.X, padx=20, pady=5)

        ttk.Label(self.top, text="Preço Unitário (R$):").pack(pady=(10, 0))
        self.preco_entry = ttk.Entry(self.top)
        self.preco_entry.pack(fill=tk.X, padx=20, pady=5)

        ttk.Label(self.top, text="Quantidade em Estoque:").pack(pady=(10, 0))
        self.quantidade_entry = ttk.Entry(self.top)
        self.quantidade_entry.pack(fill=tk.X, padx=20, pady=5)

        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=tk.X, pady=20)

        ttk.Button(btn_frame, text="Cancelar", command=self.top.destroy,
                   style='Danger.TButton').pack(side=tk.LEFT, padx=10, expand=True)
        ttk.Button(btn_frame, text="Salvar", command=self.salvar_produto,
                   style='Accent.TButton').pack(side=tk.RIGHT, padx=10, expand=True)

    def salvar_produto(self):
        nome = self.nome_entry.get().strip().upper()
        preco = self.preco_entry.get().strip()
        quantidade = self.quantidade_entry.get().strip()

        if not nome or not preco or not quantidade:
            messagebox.showerror("Erro", "Todos os campos são obrigatórios")
            return

        try:
            preco = float(preco)
            quantidade = int(quantidade)

            if preco <= 0 or quantidade < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Erro", "Preço e quantidade devem ser números válidos e positivos")
            return

        # Verificar se produto já existe
        for produto in estoque:
            if produto['Nome'] == nome:
                messagebox.showerror("Erro", "Produto já existe. Use a opção de atualizar estoque.")
                return

        codigo = gerar_codigo_produto()
        produto = {
            'Codigo': codigo,
            'Nome': nome,
            'Preco': preco,
            'Quantidade': quantidade
        }
        estoque.append(produto)
        salvar_estoque()
        messagebox.showinfo("Sucesso", f"Produto cadastrado com sucesso! Código: {codigo}")
        self.top.destroy()


class AtualizarEstoqueDialog:
    def __init__(self, parent, codigo):
        self.codigo = codigo
        self.top = tk.Toplevel(parent)
        self.top.title("Atualizar Estoque")
        self.top.geometry("400x200")

        # Encontrar produto
        self.produto = None
        for p in estoque:
            if p['Codigo'] == codigo:
                self.produto = p
                break

        if not self.produto:
            messagebox.showerror("Erro", "Produto não encontrado")
            self.top.destroy()
            return

        ttk.Label(self.top, text=f"Produto: {self.produto['Nome']}", style='Header.TLabel').pack(pady=10)
        ttk.Label(self.top, text=f"Estoque atual: {self.produto['Quantidade']}").pack()

        ttk.Label(self.top, text="Quantidade a adicionar/remover (use - para remover):").pack(pady=(10, 0))
        self.quantidade_entry = ttk.Entry(self.top)
        self.quantidade_entry.pack(fill=tk.X, padx=20, pady=5)

        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=tk.X, pady=20)

        ttk.Button(btn_frame, text="Cancelar", command=self.top.destroy,
                   style='Danger.TButton').pack(side=tk.LEFT, padx=10, expand=True)
        ttk.Button(btn_frame, text="Atualizar", command=self.atualizar_estoque,
                   style='Accent.TButton').pack(side=tk.RIGHT, padx=10, expand=True)

    def atualizar_estoque(self):
        quantidade = self.quantidade_entry.get().strip()

        if not quantidade:
            messagebox.showerror("Erro", "Digite uma quantidade")
            return

        try:
            quantidade = int(quantidade)
        except ValueError:
            messagebox.showerror("Erro", "Quantidade deve ser um número inteiro")
            return

        self.produto['Quantidade'] += quantidade
        if self.produto['Quantidade'] < 0:
            self.produto['Quantidade'] = 0
            messagebox.showwarning("Aviso", "Quantidade ajustada para zero (não pode ser negativo)")

        salvar_estoque()
        messagebox.showinfo("Sucesso", "Estoque atualizado com sucesso!")
        self.top.destroy()


# Frame de Vendas
class VendasFrame(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.carrinho = []
        self.total_venda = 0.0
        self.id_venda = str(uuid.uuid4())[:8].upper()
        self.setup_ui()

    def setup_ui(self):
        # Frame de cabeçalho
        header_frame = ttk.Frame(self)
        header_frame.pack(fill=tk.X, pady=10)

        ttk.Label(header_frame, text="Gerenciamento de Vendas", style='Header.TLabel').pack(side=tk.LEFT)

        # Frame de botões
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, pady=5)

        ttk.Button(btn_frame, text="Nova Venda", command=self.nova_venda,
                   style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Registrar Serviço", command=self.registrar_servico,
                   style='Primary.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Fechamento Diário", command=self.fechamento_diario,
                   style='Primary.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Limpar Histórico", command=self.limpar_historico,
                   style='Danger.TButton').pack(side=tk.LEFT, padx=5)

        # Notebook para abas
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Aba de Carrinho de Vendas
        self.carrinho_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.carrinho_frame, text="Carrinho de Vendas")

        # Frame de seleção de cliente
        cliente_frame = ttk.Frame(self.carrinho_frame)
        cliente_frame.pack(fill=tk.X, pady=5)

        ttk.Label(cliente_frame, text="Cliente:").pack(side=tk.LEFT)

        self.cliente_var = tk.StringVar()
        self.cliente_combobox = ttk.Combobox(cliente_frame, textvariable=self.cliente_var, state='readonly')
        self.cliente_combobox.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        self.atualizar_clientes_combobox()

        # Frame de produtos
        produto_frame = ttk.Frame(self.carrinho_frame)
        produto_frame.pack(fill=tk.X, pady=5)

        ttk.Label(produto_frame, text="Produto:").pack(side=tk.LEFT)

        self.produto_var = tk.StringVar()
        self.produto_combobox = ttk.Combobox(produto_frame, textvariable=self.produto_var)
        self.produto_combobox.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        self.atualizar_produtos_combobox()

        ttk.Label(produto_frame, text="Quantidade:").pack(side=tk.LEFT, padx=(10, 0))
        self.quantidade_entry = ttk.Entry(produto_frame, width=5)
        self.quantidade_entry.pack(side=tk.LEFT)
        self.quantidade_entry.insert(0, "1")

        ttk.Button(produto_frame, text="Adicionar", command=self.adicionar_produto_carrinho,
                   style='Primary.TButton').pack(side=tk.LEFT, padx=5)

        # Lista de itens no carrinho
        self.carrinho_tree_frame = ttk.Frame(self.carrinho_frame)
        self.carrinho_tree_frame.pack(fill=tk.BOTH, expand=True)

        self.carrinho_tree = ttk.Treeview(self.carrinho_tree_frame,
                                          columns=('Item', 'Quantidade', 'Preço Unitário', 'Subtotal'), show='headings')
        self.carrinho_tree.heading('Item', text='Item')
        self.carrinho_tree.heading('Quantidade', text='Quantidade')
        self.carrinho_tree.heading('Preço Unitário', text='Preço Unitário (R$)')
        self.carrinho_tree.heading('Subtotal', text='Subtotal (R$)')

        vsb = ttk.Scrollbar(self.carrinho_tree_frame, orient="vertical", command=self.carrinho_tree.yview)
        hsb = ttk.Scrollbar(self.carrinho_tree_frame, orient="horizontal", command=self.carrinho_tree.xview)
        self.carrinho_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.carrinho_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        self.carrinho_tree_frame.grid_rowconfigure(0, weight=1)
        self.carrinho_tree_frame.grid_columnconfigure(0, weight=1)

        # Frame de total e botões
        total_frame = ttk.Frame(self.carrinho_frame)
        total_frame.pack(fill=tk.X, pady=5)

        self.total_label = ttk.Label(total_frame, text="Total: R$ 0.00", style='Header.TLabel')
        self.total_label.pack(side=tk.LEFT)

        ttk.Button(total_frame, text="Finalizar Venda", command=self.finalizar_venda,
                   style='Accent.TButton').pack(side=tk.RIGHT, padx=5)
        ttk.Button(total_frame, text="Cancelar Venda", command=self.cancelar_venda,
                   style='Danger.TButton').pack(side=tk.RIGHT, padx=5)

        # Aba de Histórico de Vendas
        self.historico_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.historico_frame, text="Histórico de Vendas")

        # Lista de vendas
        self.historico_tree_frame = ttk.Frame(self.historico_frame)
        self.historico_tree_frame.pack(fill=tk.BOTH, expand=True)

        self.historico_tree = ttk.Treeview(self.historico_tree_frame,
                                           columns=('Tipo', 'Descrição', 'Valor', 'Cliente', 'Data', 'Qtd'),
                                           show='headings')
        self.historico_tree.heading('Tipo', text='Tipo')
        self.historico_tree.heading('Descrição', text='Descrição')
        self.historico_tree.heading('Valor', text='Valor (R$)')
        self.historico_tree.heading('Cliente', text='Cliente')
        self.historico_tree.heading('Data', text='Data/Hora')
        self.historico_tree.heading('Qtd', text='Qtd.')

        vsb = ttk.Scrollbar(self.historico_tree_frame, orient="vertical", command=self.historico_tree.yview)
        hsb = ttk.Scrollbar(self.historico_tree_frame, orient="horizontal", command=self.historico_tree.xview)
        self.historico_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.historico_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        self.historico_tree_frame.grid_rowconfigure(0, weight=1)
        self.historico_tree_frame.grid_columnconfigure(0, weight=1)

        # Carregar histórico
        self.carregar_historico()

    def atualizar_clientes_combobox(self):
        clientes_list = [f"{c['Nome']} ({c['CPF']})" for c in clientes]
        self.cliente_combobox['values'] = clientes_list
        if clientes_list:
            self.cliente_combobox.current(0)

    def atualizar_produtos_combobox(self):
        produtos_list = [f"{p['Codigo']} - {p['Nome']} (R${p['Preco']:.2f})" for p in estoque]
        self.produto_combobox['values'] = produtos_list
        if produtos_list:
            self.produto_combobox.current(0)

    def adicionar_produto_carrinho(self):
        produto_str = self.produto_var.get()
        quantidade_str = self.quantidade_entry.get()

        if not produto_str or not quantidade_str:
            messagebox.showwarning("Aviso", "Selecione um produto e informe a quantidade")
            return

        try:
            quantidade = int(quantidade_str)
            if quantidade <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Erro", "Quantidade deve ser um número inteiro positivo")
            return

        # Extrair código do produto
        codigo = produto_str.split()[0]

        # Encontrar produto no estoque
        produto = None
        for p in estoque:
            if p['Codigo'] == codigo:
                produto = p
                break

        if not produto:
            messagebox.showerror("Erro", "Produto não encontrado")
            return

        if quantidade > produto['Quantidade']:
            messagebox.showerror("Erro", f"Estoque insuficiente. Disponível: {produto['Quantidade']}")
            return

        # Adicionar ao carrinho
        valor_item = produto['Preco'] * quantidade
        self.carrinho.append({
            'tipo': 'PRODUTO',
            'descricao': produto['Nome'],
            'valor': valor_item,
            'codigo': produto['Codigo'],
            'quantidade': quantidade
        })

        # Atualizar total
        self.total_venda += valor_item
        self.total_label.config(text=f"Total: {formatar_moeda(self.total_venda)}")

        # Atualizar lista do carrinho
        self.atualizar_carrinho()

        # Reduzir estoque (temporário até finalizar venda)
        produto['Quantidade'] -= quantidade

        messagebox.showinfo("Sucesso", f"{quantidade}x {produto['Nome']} adicionado(s) ao carrinho")

    def atualizar_carrinho(self):
        for item in self.carrinho_tree.get_children():
            self.carrinho_tree.delete(item)

        for item in self.carrinho:
            self.carrinho_tree.insert('', 'end', values=(
                item['descricao'],
                item['quantidade'],
                formatar_moeda(item['valor'] / item['quantidade']),
                formatar_moeda(item['valor'])
            ))

    def finalizar_venda(self):
        if not self.carrinho:
            messagebox.showwarning("Aviso", "Carrinho vazio. Adicione itens antes de finalizar.")
            return

        cliente_str = self.cliente_var.get()
        if not cliente_str:
            messagebox.showwarning("Aviso", "Selecione um cliente para a venda")
            return

        # Extrair CPF do cliente
        cpf = cliente_str.split('(')[1].rstrip(')')

        # Encontrar cliente
        cliente = None
        for c in clientes:
            if c['CPF'] == cpf:
                cliente = c
                break

        if not cliente:
            messagebox.showerror("Erro", "Cliente não encontrado")
            return

        # Registrar venda
        data_hora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        wb, ws = carregar_planilha_vendas()

        for item in self.carrinho:
            ws.append([
                item['tipo'],
                item['descricao'],
                item['valor'],
                cliente['Nome'],
                cliente['CPF'],
                data_hora,
                item['codigo'],
                item['quantidade'],
                self.id_venda
            ])

        ajustar_largura_colunas(ws)
        wb.save(NOME_ARQUIVO_VENDAS)
        salvar_estoque()

        messagebox.showinfo("Sucesso", f"Venda finalizada com sucesso! Total: {formatar_moeda(self.total_venda)}")

        # Limpar carrinho
        self.carrinho = []
        self.total_venda = 0.0
        self.id_venda = str(uuid.uuid4())[:8].upper()
        self.total_label.config(text="Total: R$ 0.00")
        self.atualizar_carrinho()
        self.carregar_historico()

    def cancelar_venda(self):
        if not self.carrinho:
            return

        if messagebox.askyesno("Confirmar", "Cancelar venda e esvaziar carrinho?"):
            # Devolver itens ao estoque
            for item in self.carrinho:
                if item['tipo'] == 'PRODUTO':
                    for produto in estoque:
                        if produto['Codigo'] == item['codigo']:
                            produto['Quantidade'] += item['quantidade']
                            break

            self.carrinho = []
            self.total_venda = 0.0
            self.id_venda = str(uuid.uuid4())[:8].upper()
            self.total_label.config(text="Total: R$ 0.00")
            self.atualizar_carrinho()

    def nova_venda(self):
        self.notebook.select(self.carrinho_frame)

    def registrar_servico(self):
        dialog = RegistrarServicoDialog(self)
        self.wait_window(dialog.top)
        self.carregar_historico()

    def fechamento_diario(self):
        try:
            fechamento_caixa_diario()
            messagebox.showinfo("Sucesso", "Fechamento diário gerado com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao gerar fechamento: {str(e)}")

    def limpar_historico(self):
        if messagebox.askyesno("Confirmar", "Tem certeza que deseja limpar TODO o histórico de vendas?"):
            try:
                limpar_historico_vendas()
                messagebox.showinfo("Sucesso", "Histórico de vendas limpo com sucesso!")
                self.carregar_historico()
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao limpar histórico: {str(e)}")

    def carregar_historico(self):
        for item in self.historico_tree.get_children():
            self.historico_tree.delete(item)

        try:
            wb, ws = carregar_planilha_vendas()

            if ws.max_row > 1:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    self.historico_tree.insert('', 'end', values=(
                        row[0],  # Tipo
                        row[1],  # Descrição
                        formatar_moeda(row[2]),  # Valor
                        row[3],  # Cliente
                        row[5],  # Data/Hora
                        row[7]  # Quantidade
                    ))
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar histórico: {str(e)}")


# Diálogo para registrar serviço
class RegistrarServicoDialog:
    def __init__(self, parent):
        self.top = tk.Toplevel(parent)
        self.top.title("Registrar Serviço")
        self.top.geometry("500x400")

        # Frame de cliente
        cliente_frame = ttk.Frame(self.top)
        cliente_frame.pack(fill=tk.X, pady=5)

        ttk.Label(cliente_frame, text="Cliente:").pack(side=tk.LEFT)

        self.cliente_var = tk.StringVar()
        self.cliente_combobox = ttk.Combobox(cliente_frame, textvariable=self.cliente_var, state='readonly')
        self.cliente_combobox.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Carregar clientes
        clientes_list = [f"{c['Nome']} ({c['CPF']})" for c in clientes]
        self.cliente_combobox['values'] = clientes_list
        if clientes_list:
            self.cliente_combobox.current(0)

        # Frame de descrição
        descricao_frame = ttk.Frame(self.top)
        descricao_frame.pack(fill=tk.X, pady=5)

        ttk.Label(descricao_frame, text="Descrição do Serviço:").pack(side=tk.LEFT)
        self.descricao_entry = ttk.Entry(descricao_frame)
        self.descricao_entry.pack(fill=tk.X, padx=5, expand=True)

        # Frame de valor
        valor_frame = ttk.Frame(self.top)
        valor_frame.pack(fill=tk.X, pady=5)

        ttk.Label(valor_frame, text="Valor (R$):").pack(side=tk.LEFT)
        self.valor_entry = ttk.Entry(valor_frame)
        self.valor_entry.pack(side=tk.LEFT, padx=5)

        # Botões
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=tk.X, pady=20)

        ttk.Button(btn_frame, text="Cancelar", command=self.top.destroy,
                   style='Danger.TButton').pack(side=tk.LEFT, padx=10, expand=True)
        ttk.Button(btn_frame, text="Registrar", command=self.registrar_servico,
                   style='Accent.TButton').pack(side=tk.RIGHT, padx=10, expand=True)

    def registrar_servico(self):
        cliente_str = self.cliente_var.get()
        descricao = self.descricao_entry.get().strip().upper()
        valor_str = self.valor_entry.get().strip()

        if not cliente_str or not descricao or not valor_str:
            messagebox.showwarning("Aviso", "Todos os campos são obrigatórios")
            return

        try:
            valor = float(valor_str)
            if valor <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Erro", "Valor deve ser um número positivo")
            return

        # Extrair CPF do cliente
        cpf = cliente_str.split('(')[1].rstrip(')')

        # Encontrar cliente
        cliente = None
        for c in clientes:
            if c['CPF'] == cpf:
                cliente = c
                break

        if not cliente:
            messagebox.showerror("Erro", "Cliente não encontrado")
            return

        # Registrar serviço
        data_hora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        wb, ws = carregar_planilha_vendas()

        ws.append([
            "SERVIÇO",
            descricao,
            valor,
            cliente['Nome'],
            cliente['CPF'],
            data_hora,
            "N/A",
            1,
            str(uuid.uuid4())[:8].upper()
        ])

        ajustar_largura_colunas(ws)
        wb.save(NOME_ARQUIVO_VENDAS)

        messagebox.showinfo("Sucesso", f"Serviço registrado: {descricao} - Valor: {formatar_moeda(valor)}")
        self.top.destroy()


# Função principal para iniciar a aplicação
def main():
    root = tk.Tk()
    app = SistemaVendasApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()