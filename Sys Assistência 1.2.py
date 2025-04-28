from openpyxl import Workbook, load_workbook
import os

clientes = []
vendas = []

def salvar_vendas_excel():
    if not vendas:
        print('‚ö†Ô∏è Nenhuma venda registrada para salvar.')
        return

    nome_arquivo = "vendas.xlsx"

    if os.path.exists(nome_arquivo):
        wb = load_workbook(nome_arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendas"
        ws.append(["Produto", "Valor (R$)", "Cliente", "CPF"])

    # Inserir as novas vendas
    for venda in vendas:
        ws.append([
            venda['Produto'],
            venda['Valor'],
            venda['Cliente'],
            venda['CPF']
        ])

    wb.save(nome_arquivo)
    print(f'‚úÖ Vendas salvas/adicionadas com sucesso no arquivo "{nome_arquivo}"!')
    vendas.clear()
def criar_cliente():
    global clientes
    nome = input('Digite o nome do cliente: ')
    cpf = input('Digite o cpf do cliente: ')
    telefone = input('Digite o telefone do cliente: ')

    cliente = {
        'Nome': nome,
        'CPF': cpf,
        'Telefone': telefone
    }

    clientes.append(cliente)
    print('Cliente cadastrado com sucesso!')
def exibir_clientes():
    print(clientes)
def remover_cliente():
    if not clientes:
        print('‚ö†Ô∏è Nenhum cliente para deletar.')
        return

    print('\nüìã Lista de Clientes:')
    for i, cliente in enumerate(clientes, 1):
        print(f"{i}. Nome: {cliente['Nome']} | CPF: {cliente['CPF']} | Telefone: {cliente['Telefone']}")

    try:
        indice = int(input('Digite o n√∫mero do cliente que deseja deletar: '))
        if 1 <= indice <= len(clientes):
            removido = clientes.pop(indice - 1)
            print(f"‚úÖ Cliente '{removido['Nome']}' removido com sucesso.")
        else:
            print('‚ùå √çndice inv√°lido.')
    except ValueError:
        print('‚ùå Entrada inv√°lida. Digite um n√∫mero.')

def venda():
    global vendas
    if not clientes:
        print('Nenhum cliente para registrar uma venda')
        return
    print('\nLista de clientes:')
    for i, cliente in enumerate(clientes, 1):
        print(f"{i}. Nome: {cliente['Nome']} | CPF: {cliente['CPF']} | Telefone: {cliente['Telefone']}")

    try:
        indice = int(input('Digite o n√∫mero do cliente para vincular a venda'))
    except ValueError:
        print('Digite um valor v√°lido')
        return

    if 1 <= indice <= len(clientes):
        cliente_selecionado = clientes[indice - 1]
    else:
        print('Cliente inv√°lido')
        return

    produto = input('Digite o nome do produto: ').upper()
    valor = float(input('Digite o valor do produto: '))

    venda_realizada = {
    'Produto': produto,
    'Valor': valor,
    'Cliente': cliente_selecionado['Nome'],
    'CPF': cliente_selecionado['CPF']
    }
    vendas.append(venda_realizada)
    print('Venda registrada com sucesso!')

def lista_vendas():
    global vendas
    if not vendas:
        print('N√£o h√° vendas para exibir!')
        return

    print('\nLista de Vendas:')
    for i, venda_realizada in enumerate(vendas, 1):
        print(
            f"{i}. Produto: {venda_realizada['Produto']} | Valor: R${venda_realizada['Valor']:.2f} | Cliente: {venda_realizada['Cliente']} | CPF: {venda_realizada['CPF']}")
while True:
    print('BEM VINDO!\n1- Criar Cliente\n2- Exibir lista de cliente\n3- Remover Cliente\n4- Registrar venda\n5-Exibir lista de vendas\n6-Salvar o registro de vendas')
    escolha = input('Digite sua escolha: ')
    match escolha:
        case '1':
            criar_cliente()
        case '2':
            exibir_clientes()
        case '3':
            remover_cliente()
        case '4':
            venda()
        case '5':
            lista_vendas()
        case '6':
            salvar_excel()
        case '_':
            print('Op√ß√£o inv√°lida!')